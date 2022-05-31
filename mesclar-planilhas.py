from openpyxl import Workbook, load_workbook

nomeArquivo1 = input('Digite o nome da planilha enviada pela transportadora: ')
nomeArquivo2 = input('Digite o nome da planilha principal (Controle): ')

planilha1 = load_workbook(nomeArquivo1 + ".xlsx")

aba_ativa_p1 = planilha1.active

planilha2 = load_workbook(nomeArquivo2 + ".xlsx")

print("Executando...")

aba_ativa_p2 = planilha2["01"]

for celula in aba_ativa_p1["C"]:
    linha = celula.row
    if linha != 1:
        for celula2 in aba_ativa_p2["D"]:
            linha2 = celula2.row
            if linha2 != 1:
                if aba_ativa_p1[f"D{linha}"].value == aba_ativa_p2[f"D{linha2}"].value:
                    aba_ativa_p2[f"J{linha2}"] = aba_ativa_p1[f"G{linha}"].value

print("Planilha criada com sucesso!!!")

novo = input('Digite o nome da nova planilha: ')

planilha2.save(novo + ".xlsx")